#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Optimize weekly employee->schedule assignments with preferences,
availability (all-zero COST column = unavailable), days off allowed,
and weekly MAX HOURS per employee.

Excel layout:
- Each day sheet (Mon..Sun)
  * COST:   B3:U22 (20x20)
  * PREF:   W3:AP22 (20x20)
  * HOURS:  B24:U24 (1x20)  <-- NEW (hours per schedule A0..T0)
  * X (out):B26:U45 (20x20) decisions written by script

- Weekly sheet
  * λ (pref weight): E2
  * Max Shifts:      C6:C25 (<= weekly shifts)
  * Max Hours:       D6:D25 (<= weekly hours)  <-- NEW

Usage:
  py -3.11 optimize_schedules_with_sanity.py --input StaffScheduler.xlsx --output StaffScheduler_Solved.xlsx
"""

import argparse
import os
import numpy as np
from openpyxl import load_workbook

DAYS = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
N = 20
SCALE = 1000  # integerize objective for CP-SAT

# -----------------------------
# I/O
# -----------------------------
def read_cost_pref_hours_caps(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    costs, prefs, hours = {}, {}, {}

    for d in DAYS:
        ws = wb[d]
        # COST: B3:U22
        C = np.zeros((N, N), dtype=float)
        for i in range(N):
            for j in range(N):
                v = ws.cell(row=3+i, column=2+j).value
                C[i, j] = float(v or 0.0)
        costs[d] = C

        # PREF: W3:AP22
        P = np.zeros((N, N), dtype=float)
        for i in range(N):
            for j in range(N):
                v = ws.cell(row=3+i, column=23+j).value
                P[i, j] = float(v or 0.0)
        prefs[d] = P

        # HOURS per schedule: B24:U24  (vector length 20)
        h = np.zeros(N, dtype=float)
        for s in range(N):
            v = ws.cell(row=24, column=2+s).value
            h[s] = float(v or 0.0)
        hours[d] = h

    ws_w = wb["Weekly"]
    lam = float(ws_w["E2"].value or 1.0)

    # Weekly shift caps C6:C25
    shift_caps = []
    for r in range(6, 26):
        v = ws_w.cell(row=r, column=3).value
        shift_caps.append(int(v) if v is not None else 7)  # <= shifts per week

    # Weekly hour caps F6:F25 (optional; 0/blank means no hour cap)
    hour_caps = []
    for r in range(6, 26):
        v = ws_w.cell(row=r, column=6).value
        hour_caps.append(float(v or 0.0))  # <= hours per week; 0 => no cap

    return costs, prefs, hours, lam, shift_caps, hour_caps

def write_solution(xlsx_in: str, sol, objective: float, xlsx_out: str, costs, hours, hour_caps):
    print("[Banner] write_solution() called")
    print(f"[Banner] Input:  {os.path.abspath(xlsx_in)}")
    print(f"[Banner] Output: {os.path.abspath(xlsx_out)}")

    wb = load_workbook(xlsx_in, data_only=False)
    day_idx = {d:i for i,d in enumerate(DAYS)}

    # 1) decisions to day sheets
    for d in DAYS:
        ws = wb[d]
        di = day_idx[d]
        # clear B26:U45
        for i in range(N):
            for j in range(N):
                ws.cell(row=26+i, column=2+j).value = 0
        # set 1's
        for e in range(N):
            for s in range(N):
                if sol[(e,s,di)] == 1:
                    ws.cell(row=26+e, column=2+s).value = 1

    # 2) weekly objective + (optional) total hours per employee
    ws_w = wb["Weekly"]
    ws_w["A3"] = "Solved Objective (from Python)"
    ws_w["B3"] = float(objective)

    # compute weekly hours per employee for display
    weekly_hours = [0.0]*N
    for e in range(N):
        total = 0.0
        for d, day in enumerate(DAYS):
            hvec = hours[day]
            for s in range(N):
                if sol[(e,s,d)] == 1:
                    total += hvec[s]
        weekly_hours[e] = total
        # write to Weekly column E (optional; easy to see next to cap in D)
        ws_w.cell(row=6+e, column=5).value = total  # E6:E25 = actual hours

    ws_w["D5"] = "Max Hours (cap)"
    ws_w["E5"] = "Actual Hours (from Python)"

    # 3) Sanity sheet (usage, availability, employee counts, hour totals)
    if "Sanity" in wb.sheetnames:
        wb.remove(wb["Sanity"])
    ws_sanity = wb.create_sheet("Sanity")

    # headers
    ws_sanity["A1"] = "Schedule usage per day (sum over employees for each schedule)"
    for s in range(N):
        ws_sanity.cell(row=2, column=2+s).value = f"S{s}"  # S0..S19 -> A0..T0

    base_av = 10
    ws_sanity.cell(row=base_av, column=1).value = "Availability per day (1=available, 0=unavailable; availability = any non-zero COST in column)"

    base_emp = 20
    ws_sanity.cell(row=base_emp, column=1).value = "Employee assignment count per day (<=1 expected)"
    for e in range(N):
        ws_sanity.cell(row=base_emp+1, column=2+e).value = f"E{e+1}"

    base_hours = 32
    ws_sanity.cell(row=base_hours, column=1).value = "Weekly hours per employee (cap vs actual)"
    ws_sanity.cell(row=base_hours, column=2).value = "Cap"
    ws_sanity.cell(row=base_hours, column=3).value = "Actual"
    for e in range(N):
        ws_sanity.cell(row=base_hours+1+e, column=1).value = f"E{e+1}"
        ws_sanity.cell(row=base_hours+1+e, column=2).value = hour_caps[e]
        ws_sanity.cell(row=base_hours+1+e, column=3).value = weekly_hours[e]

    violations = []  # (day, s) if picked schedule that is unavailable

    for d, day in enumerate(DAYS):
        # usage row
        ws_sanity.cell(row=3+d, column=1).value = day
        # availability row
        ws_sanity.cell(row=base_av+1+d, column=1).value = day

        C = costs[day]
        avail_col = [0 if np.all(C[:, s] == 0.0) else 1 for s in range(N)]

        for s in range(N):
            # usage count
            usage = sum(sol[(e,s,d)] for e in range(N))
            ws_sanity.cell(row=3+d, column=2+s).value = usage
            # availability
            ws_sanity.cell(row=base_av+1+d, column=2+s).value = avail_col[s]
            if avail_col[s] == 0 and usage != 0:
                violations.append((day, s))

        # per-employee count (0/1)
        for e in range(N):
            cnt = sum(sol[(e,s,d)] for s in range(N))
            ws_sanity.cell(row=base_emp+2+d, column=2+e).value = cnt

    base_viol = base_hours + 25
    ws_sanity.cell(row=base_viol, column=1).value = "Violations (picked schedule with all-zero COST column)"
    if violations:
        for idx, (day, s) in enumerate(violations, start=1):
            ws_sanity.cell(row=base_viol+idx, column=1).value = f"{day} : S{s}"
    else:
        ws_sanity.cell(row=base_viol+1, column=1).value = "None"

    print(f"[Banner] Sheets before save: {wb.sheetnames}")
    wb.save(xlsx_out)
    print(f"[Banner] Saved to: {os.path.abspath(xlsx_out)}")

# -----------------------------
# Solver
# -----------------------------
def solve_cpsat(costs, prefs, hours, lam, shift_caps, hour_caps, max_time=60, workers=8):
    from ortools.sat.python import cp_model
    model = cp_model.CpModel()

    # Decision vars
    x = {(e,s,d): model.NewBoolVar(f"x_e{e}_s{s}_d{d}")
         for e in range(N) for s in range(N) for d in range(7)}

    # Objective: min sum_d,e,s (C - lam*P) * x  (scaled)
    terms = []
    for d, day in enumerate(DAYS):
        C = costs[day]; P = prefs[day]
        for e in range(N):
            for s in range(N):
                coef = C[e, s] - lam * P[e, s]
                terms.append((int(round(coef * SCALE)), x[(e, s, d)]))
    model.Minimize(sum(c * var for c, var in terms))

    # Constraints per day (availability + days off)
    for d, day in enumerate(DAYS):
        # employee at most one schedule per day
        for e in range(N):
            model.Add(sum(x[(e, s, d)] for s in range(N)) <= 1)

        # schedule availability: all-zero COST column => forbid; else exactly once
        C = costs[day]
        for s in range(N):
            col = C[:, s]
            if np.all(col == 0.0):
                for e in range(N):
                    model.Add(x[(e, s, d)] == 0)
            else:
                model.Add(sum(x[(e, s, d)] for e in range(N)) == 1)

    # Weekly shift caps (if you use them)
    for e in range(N):
        model.Add(sum(x[(e, s, d)] for d in range(7) for s in range(N)) <= int(shift_caps[e]))

    # Weekly hour caps (NEW): sum_d,s hours[d][s] * x[e,s,d] <= hour_caps[e]  (skip if cap <= 0)
    for e in range(N):
        cap = float(hour_caps[e])
        if cap > 0:
            model.Add(
                sum(int(round(hours[day][s] * SCALE)) * x[(e, s, d)]
                    for d, day in enumerate(DAYS) for s in range(N))
                <= int(round(cap * SCALE))
            )

    # Solve
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(max_time)
    solver.parameters.num_search_workers = int(workers)
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(
            "No feasible solution found. Check availability, Weekly caps, and Max Hours."
        )

    # Extract solution
    sol = {(e, s, d): int(solver.Value(x[(e, s, d)]))
           for e in range(N) for s in range(N) for d in range(7)}

    # Unscaled objective
    obj = 0.0
    for d, day in enumerate(DAYS):
        C = costs[day]; P = prefs[day]
        for e in range(N):
            for s in range(N):
                if sol[(e, s, d)] == 1:
                    obj += C[e, s] - lam * P[e, s]
    return sol, obj

# -----------------------------
# Main
# -----------------------------
def main():
    print(f"[Banner] Running: {os.path.abspath(__file__)}")

    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--max_time", type=float, default=60.0)
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    costs, prefs, hours, lam, shift_caps, hour_caps = read_cost_pref_hours_caps(args.input)

    # Quick precheck: available schedules per day must be <= N (days off allowed)
    issues = []
    for day in DAYS:
        C = costs[day]
        available = sum(0 if np.all(C[:, s] == 0.0) else 1 for s in range(N))
        if available > N:
            issues.append((day, available))
    if issues:
        print("[Precheck] Days where available schedules > number of employees (N) -> infeasible with '==1 per available schedule'.")
        for day, k in issues:
            print(f"  - {day}: available schedules = {k}, N = {N}")

    sol, obj = solve_cpsat(costs, prefs, hours, lam, shift_caps, hour_caps,
                           max_time=args.max_time, workers=args.workers)
    write_solution(args.input, sol, obj, args.output, costs, hours, hour_caps)
    print(f"Solved. Objective={obj:.2f}. Wrote {args.output}")

if __name__ == "__main__":
    main()
